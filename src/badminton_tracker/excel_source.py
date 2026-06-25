"""Read reference data (friend nicknames, columns) from the source workbook."""

from __future__ import annotations

import openpyxl

from .config import SOURCE_XLSX

# Column order of the source "Data" sheet — the target schema for outputs.
DATA_COLUMNS = [
    "Date",
    "Tournament",
    "Category",
    "Level",
    "Round",
    "Player 1",
    "Player 2",
    "Opponent 1",
    "Opponent 2",
    "Result",
    "Set 1 OWN",
    "Set 1 OPP",
    "Set 2 OWN",
    "Set 2 OPP",
    "Set 3 OWN",
    "Set 3 OPP",
]


def friend_names() -> list[str]:
    """Distinct names that appear in the Player 1 / Player 2 columns of the log."""
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb["Data"]
    names: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7, values_only=True):
        for cell in row:
            if cell and str(cell).strip() and str(cell).strip() not in seen:
                seen.add(str(cell).strip())
                names.append(str(cell).strip())
    wb.close()
    return names


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def read_data_matches() -> list[dict]:
    """Read the source workbook's Data sheet into the internal match-dict format."""
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb["Data"]
    matches: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:  # no tournament -> empty row
            continue
        date = row[0]
        matches.append(
            {
                "date": date.date().isoformat() if hasattr(date, "date") else _s(date),
                "tournament": _s(row[1]),
                "category": _s(row[2]),
                "level": _s(row[3]),
                "round": _s(row[4]),
                "player_1": _s(row[5]),
                "player_2": _s(row[6]),
                "opponent_1": _s(row[7]),
                "opponent_2": _s(row[8]),
                "result": _s(row[9]).upper(),
                "set_1_own": row[10],
                "set_1_opp": row[11],
                "set_2_own": row[12],
                "set_2_opp": row[13],
                "set_3_own": row[14],
                "set_3_opp": row[15],
            }
        )
    wb.close()
    return matches
