"""Security + format validation for maintainer-supplied files.

Two untrusted inputs reach the server from the web UI:

* an uploaded `.xlsx` workbook (Santeri replacing the match log), and
* edited alias rows (friends mapping names → nicknames).

Everything here is defensive: validate *before* anything touches disk or git.
Nothing in this module trusts a filename, a content-type header, or row counts
claimed by the client.
"""

from __future__ import annotations

import io
import zipfile

import openpyxl

from .config import MAX_UPLOAD_BYTES
from .excel_source import DATA_COLUMNS

# A real .xlsx is a ZIP container; first bytes are the local-file-header magic.
_ZIP_MAGIC = b"PK\x03\x04"
_EMPTY_ZIP_MAGIC = b"PK\x05\x06"  # an empty archive

MAX_ROWS = 50_000  # zip-bomb / runaway guard for the Data sheet
MAX_FIELD_LEN = 200  # any single alias/name cell
MAX_ALIAS_ROWS = 5_000


class ValidationError(ValueError):
    """Raised when an upload or edit fails a security/format check.

    `messages` is a list of human-readable problems to show the maintainer.
    """

    def __init__(self, messages: list[str]):
        self.messages = messages
        super().__init__("; ".join(messages))


def _has_control_chars(s: str) -> bool:
    # Reject NUL and other C0 controls (except tab) that have no business in a
    # name/nickname and can be used for log/CSV injection or display spoofing.
    return any(ord(c) < 32 and c != "\t" for c in s)


# ── Excel upload ───────────────────────────────────────────────────────────


def validate_workbook_bytes(raw: bytes) -> dict:
    """Validate raw uploaded bytes as the match-log workbook.

    Returns a small summary dict on success; raises ValidationError otherwise.
    """
    errors: list[str] = []

    if not raw:
        raise ValidationError(["The uploaded file is empty."])
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValidationError(
            [f"File is {len(raw):,} bytes; the limit is {MAX_UPLOAD_BYTES:,}."]
        )
    if not (raw.startswith(_ZIP_MAGIC) or raw.startswith(_EMPTY_ZIP_MAGIC)):
        raise ValidationError(
            ["This is not a real .xlsx file (bad signature). Save as Excel Workbook (.xlsx)."]
        )

    # Inspect the zip directory without trusting the extension. Refuse macro-
    # enabled content and obviously malformed archives.
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            names = zf.namelist()
            if any(n.endswith("vbaProject.bin") for n in names):
                errors.append("Macro-enabled workbooks are not allowed. Save as plain .xlsx.")
            if "xl/workbook.xml" not in names:
                errors.append("File does not look like an Excel workbook.")
    except zipfile.BadZipFile:
        raise ValidationError(["The file is corrupt or not a valid .xlsx archive."]) from None

    if errors:
        raise ValidationError(errors)

    # Structural validation: required sheet, exact column schema, sane rows.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — surface any openpyxl failure to the user
        raise ValidationError([f"Could not open the workbook: {exc}"]) from None

    try:
        if "Data" not in wb.sheetnames:
            raise ValidationError(
                [f"Missing required sheet 'Data'. Sheets found: {', '.join(wb.sheetnames)}."]
            )
        ws = wb["Data"]
        header = [
            (str(c).strip() if c is not None else "")
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ]
        # Header must start with the expected schema (trailing extra cols tolerated).
        if header[: len(DATA_COLUMNS)] != DATA_COLUMNS:
            raise ValidationError(
                [
                    "The 'Data' sheet header does not match the expected columns.",
                    f"Expected: {', '.join(DATA_COLUMNS)}",
                    f"Found:    {', '.join(header) or '(empty)'}",
                ]
            )

        rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:  # has a tournament → a real row
                rows += 1
            if rows > MAX_ROWS:
                raise ValidationError([f"Too many rows (> {MAX_ROWS:,})."])
        if rows == 0:
            raise ValidationError(["The 'Data' sheet has no match rows."])
    finally:
        wb.close()

    return {"matches": rows, "bytes": len(raw)}


# ── Alias edits ────────────────────────────────────────────────────────────

ALIAS_FIELDS = ["name", "display", "notes"]


def validate_alias_rows(rows: list[dict]) -> list[dict]:
    """Validate + normalise alias rows from the web editor.

    Returns cleaned rows (stripped, deduped order preserved) or raises.
    """
    errors: list[str] = []
    if not isinstance(rows, list):
        raise ValidationError(["Malformed payload: expected a list of rows."])
    if len(rows) > MAX_ALIAS_ROWS:
        raise ValidationError([f"Too many rows (> {MAX_ALIAS_ROWS:,})."])

    cleaned: list[dict] = []
    seen: set[str] = set()
    for i, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            errors.append(f"Row {i}: not an object.")
            continue
        name = str(row.get("name", "")).strip()
        display = str(row.get("display", "")).strip()
        notes = str(row.get("notes", "")).strip()
        if not name:
            # Silently drop fully-blank rows; flag rows with content but no name.
            if display or notes:
                errors.append(f"Row {i}: has a nickname/notes but no name.")
            continue
        for field, val in (("name", name), ("display", display), ("notes", notes)):
            if len(val) > MAX_FIELD_LEN:
                errors.append(f"Row {i}: {field} is too long (> {MAX_FIELD_LEN} chars).")
            if _has_control_chars(val):
                errors.append(f"Row {i}: {field} contains invalid control characters.")
        key = name.casefold()
        if key in seen:
            errors.append(f"Row {i}: duplicate name '{name}'.")
            continue
        seen.add(key)
        cleaned.append({"name": name, "display": display, "notes": notes})

    if errors:
        raise ValidationError(errors)
    return cleaned
