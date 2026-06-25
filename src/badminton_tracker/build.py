"""Write outputs: matches.csv and a standalone .xlsx with Python-computed stats."""

from __future__ import annotations

import csv
import datetime as dt
import re

import openpyxl
from openpyxl.utils import get_column_letter

from .config import MATCHES_MIRROR_CSV, OUT_DIR
from .excel_source import DATA_COLUMNS, read_data_matches
from .stats import player_stats, tournament_stats

MATCHES_CSV = OUT_DIR / "matches.csv"
OUTPUT_XLSX = OUT_DIR / "Badminton Bro Tournament Log - generated.xlsx"

_DATE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")

# Normalise round labels (Finnish + English) to the source workbook's short codes.
# Order matters: more specific keys must precede substrings of themselves
# ("quarterfinal"/"semifinal" before "final"; "puolivälierä" before "välierä").
_ROUND_MAP = {
    "quarterfinal": "QF",
    "quarter-final": "QF",
    "puolivälierä": "QF",
    "semifinal": "SF",
    "semi-final": "SF",
    "välierä": "SF",
    "final": "F",
    "loppuottelu": "F",
    "pool": "Pool",
    "lohko": "Pool",
    "group": "Pool",
}


def normalize_date(s: str | None) -> dt.date | None:
    if not s:
        return None
    m = _DATE_RE.search(s)
    if not m:
        return None
    day, month, year = (int(x) for x in m.groups())
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def normalize_round(s: str | None) -> str:
    if not s:
        return ""
    low = s.strip().lower()
    for key, code in _ROUND_MAP.items():
        if key in low:
            return code
    m = re.search(r"(?:round of|kierros)\s*(\d+)", low)
    if m:
        return m.group(1)
    m = re.search(r"\b(\d{1,3})\b", low)
    if m:
        return m.group(1)
    return s.strip()


def to_data_row(m: dict) -> dict:
    return {
        "Date": normalize_date(m["date"]),
        "Tournament": m["tournament"],
        "Category": m["category"],
        "Level": m["level"],
        "Round": normalize_round(m["round"]),
        "Player 1": m["player_1"],
        "Player 2": m["player_2"],
        "Opponent 1": m["opponent_1"],
        "Opponent 2": m["opponent_2"],
        "Result": m["result"],
        "Set 1 OWN": m["set_1_own"],
        "Set 1 OPP": m["set_1_opp"],
        "Set 2 OWN": m["set_2_own"],
        "Set 2 OPP": m["set_2_opp"],
        "Set 3 OWN": m["set_3_own"],
        "Set 3 OPP": m["set_3_opp"],
    }


def _sort_key(row: dict):
    return (row["Date"] or dt.date.min, row["Tournament"] or "", str(row["Category"]))


def write_csv(rows: list[dict]) -> None:
    with open(MATCHES_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=DATA_COLUMNS)
        writer.writeheader()
        for r in rows:
            out = dict(r)
            out["Date"] = r["Date"].isoformat() if r["Date"] else ""
            writer.writerow(out)


def write_matches_mirror() -> int:
    """Dump the workbook's Data sheet to a stable CSV beside it in data/.

    Committed alongside the binary .xlsx so `git -C data diff matches_mirror.csv`
    shows exactly which match rows changed between Excel uploads.
    """
    rows = sorted((to_data_row(m) for m in read_data_matches()), key=_sort_key)
    with open(MATCHES_MIRROR_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=DATA_COLUMNS)
        writer.writeheader()
        for r in rows:
            out = dict(r)
            out["Date"] = r["Date"].isoformat() if r["Date"] else ""
            writer.writerow(out)
    return len(rows)


def _write_sheet(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r.get(h, ""))) for r in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 40)
    ws.freeze_panes = "A2"


def write_xlsx(data_rows: list[dict], pstats: list[dict], tstats: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    _write_sheet(ws, DATA_COLUMNS, data_rows)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if isinstance(cell.value, dt.date):
                cell.number_format = "yyyy-mm-dd"

    p_headers = [
        "player", "games", "wins", "losses", "win_ratio",
        "sets_won", "sets_lost", "points_for", "points_against",
        "points_per_set", "third_set_ratio",
    ]
    _write_sheet(wb.create_sheet("Player statistics"), p_headers, pstats)

    t_headers = ["player", "tournament", *p_headers[1:]]
    _write_sheet(wb.create_sheet("Tournament statistics"), t_headers, tstats)

    wb.save(OUTPUT_XLSX)


def build(matches: list[dict], roster: list[dict]) -> None:
    data_rows = sorted((to_data_row(m) for m in matches), key=_sort_key)
    pstats = player_stats(matches, roster)
    tstats = tournament_stats(matches, roster)
    write_csv(data_rows)
    write_xlsx(data_rows, pstats, tstats)
    print(f"Wrote {len(data_rows)} matches -> {MATCHES_CSV}")
    print(f"Wrote workbook ({len(pstats)} players) -> {OUTPUT_XLSX}")
